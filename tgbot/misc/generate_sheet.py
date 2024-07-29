import io
import string

import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border


async def generate_sheet(sheet, column_name, list_statistic):
    font_title = Font(name='Arial', size=13, bold=True)
    font_all = Font(name='Arial', size=13)
    text_style = Alignment(horizontal='center', vertical='top', wrapText=True)
    thin_border = Side(border_style="thin")
    for i, column_name in enumerate(column_name, start=1):
        sheet.cell(row=1, column=i, value=column_name).font = font_title
        sheet.cell(row=1, column=i).border = Border(top=thin_border, bottom=thin_border, left=thin_border,
                                                    right=thin_border)
        sheet.column_dimensions[string.ascii_uppercase[i - 1]].width = 25
    for row_number, statistic in enumerate(list_statistic, start=2):
        for column_number, value in enumerate(statistic):
            sheet.cell(row=row_number, column=column_number + 1, value=value).font = font_all
            sheet.cell(row=row_number, column=column_number + 1).alignment = text_style
            sheet.cell(row=row_number, column=column_number + 1).border = Border(top=thin_border, bottom=thin_border,
                                                                                 left=thin_border, right=thin_border)
    sheet.freeze_panes = 'A2'


async def get_app_statistics(queryset):
    column_name = ['ID', 'Текст', 'Статус', 'Заявитель', 'Дата создания', 'Дата выполнения', 'Оценка']
    list_statistic = []
    for obj in queryset:
        completed_time = ''
        if obj.completed_time:
            completed_time = obj.completed_time.strftime("%d.%m.%Y %H:%M")
        list_statistic.append([
            obj.id, obj.text, obj.get_status_display(), obj.user.fcs, obj.created.strftime("%d.%m.%Y %H:%M"),
            completed_time, obj.grade
        ])
    return column_name, list_statistic


async def generate_statistic(queryset):
    file = io.BytesIO()

    book = openpyxl.Workbook()
    for sheet_name in book.sheetnames:
        sheet = book.get_sheet_by_name(sheet_name)
        book.remove_sheet(sheet)

    app_sheet = book.create_sheet('Заявки')

    column_name, list_statistic = await get_app_statistics(queryset)

    await generate_sheet(app_sheet, column_name, list_statistic)

    book.save(file)
    file.seek(0)
    return file
